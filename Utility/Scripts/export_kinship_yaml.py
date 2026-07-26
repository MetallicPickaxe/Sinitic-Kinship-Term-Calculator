import sys
import json
from pathlib import Path
from openpyxl import load_workbook
import yaml

if len(sys.argv) < 3:
    print("Usage: python export_kinship_yaml.py <excel_path> <output_yaml>")
    sys.exit(1)

excel_path = Path(sys.argv[1])
output_path = Path(sys.argv[2])

wb = load_workbook(excel_path, data_only=True)
ws = wb.active
headers = [cell.value for cell in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    record = {header: value for header, value in zip(headers, row)}
    rows.append(record)

output_path.parent.mkdir(parents=True, exist_ok=True)
with output_path.open('w', encoding='utf-8') as fh:
    yaml.safe_dump(rows, fh, allow_unicode=True, sort_keys=False)

print(f'Exported {len(rows)} entries to {output_path}')
