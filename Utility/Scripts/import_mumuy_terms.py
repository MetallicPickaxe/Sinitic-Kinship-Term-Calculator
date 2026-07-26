import json
import sys
from pathlib import Path
from openpyxl import load_workbook

# Repository root is derived from the script location (two levels above Utility/Scripts/)
# rather than hard-coding anyone's local drive layout.
REPO_ROOT = Path(__file__).resolve().parents[2]
MODE_MAP_PATH = REPO_ROOT / "Utility" / "MumuyAlgorithm" / "Data" / "mode-map.json"
HEADERS = ["relation_id","selector","symbol_path","generation","lateral_type","side","gender","spouse_parity","zh_hans","zh_hant","en","aliases","notes"]

TOKEN_MAP = {
    'f': [['F']],
    'm': [['M']],
    's': [['S']],
    'd': [['D']],
    'ob': [['OB']],
    'lb': [['YB']],
    'os': [['OS']],
    'ls': [['YS']],
    'xb': [['OB'], ['YB']],
    'xs': [['OS'], ['YS']],
    'w': [['SP']],
    'h': [['SP']],
    'sp': [['SP']],
}

GEN_DELTA = {
    'F': 1,
    'M': 1,
    'S': -1,
    'D': -1
}

UNSUPPORTED_TOKENS = set()


def split_options(content: str):
    options = []
    current = []
    depth = 0
    buffer = ''
    for ch in content:
        if ch == '[':
            depth += 1
        elif ch == ']':
            depth -= 1
        if ch == '|' and depth == 0:
            options.append(buffer)
            buffer = ''
        else:
            buffer += ch
    if buffer:
        options.append(buffer)
    return [opt.strip() for opt in options if opt.strip()]


def parse_sequence(raw: str):
    tokens = []
    i = 0
    while i < len(raw):
        ch = raw[i]
        if ch == '[':
            depth = 1
            j = i + 1
            while j < len(raw) and depth > 0:
                if raw[j] == '[':
                    depth += 1
                elif raw[j] == ']':
                    depth -= 1
                j += 1
            content = raw[i + 1:j - 1]
            option_strings = split_options(content)
            option_sequences = []
            for opt in option_strings:
                parts = [part.strip() for part in opt.split(',') if part.strip()]
                option_sequences.append(parts)
            tokens.append(option_sequences)
            i = j
        elif ch == ',':
            i += 1
        elif ch == ' ':
            i += 1
        else:
            start = i
            while i < len(raw) and raw[i] not in ',[]':
                i += 1
            token = raw[start:i].strip()
            if token:
                tokens.append(token)
    sequences = [[]]
    for part in tokens:
        if isinstance(part, str):
            if not part:
                continue
            for idx in range(len(sequences)):
                sequences[idx] = sequences[idx] + [part]
        else:
            new_sequences = []
            for seq in sequences:
                for option in part:
                    new_sequences.append(seq + option)
            sequences = new_sequences
    return sequences


def normalize_token(token: str):
    token = token.strip()
    token = token.replace('\u200b', '')
    if not token:
        return ''
    if '&' in token:
        token = token.split('&', 1)[0]
    token = token.replace("'", '')
    return token


def expand_to_symbols(sequence):
    sequences = [[]]
    for raw in sequence:
        token = normalize_token(raw)
        if not token:
            continue
        token = token.lower()
        if token not in TOKEN_MAP:
            UNSUPPORTED_TOKENS.add(token)
            return []
        options = TOKEN_MAP[token]
        new_sequences = []
        for seq in sequences:
            for option in options:
                new_sequences.append(seq + option)
        sequences = new_sequences
    return sequences


def compute_generation(symbols):
    gen = 0
    for symbol in symbols:
        gen += GEN_DELTA.get(symbol, 0)
    return gen


def compute_spouse_parity(symbols):
    return sum(1 for symbol in symbols if symbol == 'SP')


def load_mode_map():
    with MODE_MAP_PATH.open('r', encoding='utf-8') as fh:
        text = fh.read()
        if text.startswith('{'):
            return json.loads(text)
        import ast
        return ast.literal_eval(text)


def generate_entries():
    data = load_mode_map()
    entries = {}
    for key, names in data.items():
        if not names:
            continue
        zh_hans = names[0]
        base_selector = key
        raw_sequences = parse_sequence(key)
        for raw_sequence in raw_sequences:
            symbol_sequences = expand_to_symbols(raw_sequence)
            for symbols in symbol_sequences:
                if not symbols:
                    continue
                symbol_path = '.'.join(symbols)
                if symbol_path in entries:
                    continue
                selector_lower = ','.join(raw_sequence)
                entry = {
                    'relation_id': f"auto-{symbol_path.replace('.', '-')}",
                    'selector': selector_lower,
                    'symbol_path': symbol_path,
                    'generation': compute_generation(symbols),
                    'lateral_type': 'auto',
                    'side': 'auto',
                    'gender': 'unknown',
                    'spouse_parity': compute_spouse_parity(symbols),
                    'zh_hans': zh_hans,
                    'zh_hant': zh_hans,
                    'en': zh_hans,
                    'aliases': ';'.join(names[1:]) if len(names) > 1 else '',
                    'notes': 'Imported from Mumuy mode-map'
                }
                entries[symbol_path] = entry
    return entries


def merge_into_excel(excel_path: Path, new_entries: dict[str, dict]):
    wb = load_workbook(excel_path)
    ws = wb.active
    existing = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        data = {header: value for header, value in zip(HEADERS, row)}
        symbol_path = data.get('symbol_path')
        if symbol_path:
            existing[symbol_path] = True
    appended = 0
    for symbol_path, entry in new_entries.items():
        if symbol_path in existing:
            continue
        ws.append([entry.get(header, '') for header in HEADERS])
        appended += 1
    wb.save(excel_path)
    return appended


def main():
    if len(sys.argv) < 2:
        print('Usage: python import_mumuy_terms.py <excel_path>')
        return
    excel_path = Path(sys.argv[1])
    entries = generate_entries()
    appended = merge_into_excel(excel_path, entries)
    print(f'Appended {appended} new entries to {excel_path}')
    if UNSUPPORTED_TOKENS:
        print(f'Unsupported tokens encountered: {sorted(UNSUPPORTED_TOKENS)}')

if __name__ == '__main__':
    main()
